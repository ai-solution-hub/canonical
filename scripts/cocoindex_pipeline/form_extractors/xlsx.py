"""XLSX form extractor for ID-52 Path B (PRODUCT Inv-2).

Implements ``extract(raw_bytes, filename) -> ExtractedForm`` for blank
XLSX forms using ``openpyxl==3.1.5`` (pinned in ``requirements.txt``).

Design notes
------------

The extractor walks every worksheet in the workbook. For each sheet it
detects two table archetypes observed in the Path-B corpus:

* **EFA scoring-matrix** — ``Ref | Criteria | (blank) | Weighting |
  Score | Weighted Score | Remarks`` (RESEARCH §2.2; verified on
  ``evaluation-matrix-itt-vol8.xlsx``). Section-header rows have the
  Part name in the ``Criteria`` column and a weighting in column 4
  but no ``Ref``; question rows carry a ``Ref`` (e.g. ``2.1``, ``5.3``)
  and the question prose in the ``Criteria`` column.
* **EFA compliance-check** — ``Section | … | Page Limit | Page Count |
  Remarks`` (rows 4-10 on each Bidder sheet). The rows list Part names
  rather than questions, so no field is emitted from this sub-table:
  treating them as fields would clone the Part headers as content and
  inflate the dedup count. The ``Page Limit`` column is a
  ``word_limit`` substrate that the spec maps when populated — in
  this fixture it is blank on every row, so the mapping is exercised
  by the scoring-matrix Schema-B handler.
* **CSP letter-keyed preamble** — column B carries a letter key
  (``A``, ``B``, ``C``, …) and column C carries the section title
  followed (after a newline) by the question prose. Some rows have
  no letter key in column B but carry a sub-key (``B1``, ``B2``, …)
  embedded in the column C prose; these continue the prior letter's
  section.
* **CSP numbered principles** — column B carries ``PRINCIPLE N``
  (possibly wrapped over two lines in the source), column C carries
  the principle title + prose, column D carries the NCSC URL.
  Sub-rows like ``2.1``, ``9.1`` appear as numbered prefixes in
  column C and continue the prior principle's section.

Per Inv-13 the extractor performs **per-form dedup** keyed on
``(section_name, normalise(question_text))`` after walking all sheets:
this collapses the ``Bidder 1`` ≡ ``Bidder 2`` repetition in the EFA
matrix to N fields rather than 2N. The catalogue-level reuse case is
Path-C's concern (TECH §2.3).

References:
- ``docs/specs/form-extraction/PRODUCT.md`` Inv-2, Inv-7..Inv-14, Inv-17.
- ``docs/specs/form-extraction/TECH.md`` §2.2 (Pydantic shape), §2.3
  (dedup), §2.5a (mandatory flag), §2.6 (Migration M1 columns).
- ``docs/specs/form-extraction/PLAN.md`` §{52.10} (acceptance criteria).
"""

from __future__ import annotations

import io
import json
import logging
import re
from typing import Any

import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from scripts.cocoindex_pipeline.form_extractors.shared import (
    ExtractedField,
    ExtractedForm,
    FormExtractionError,
    FormMetadata,
)

__all__ = ["NO_ARCHETYPE_REASON", "extract"]

_logger = logging.getLogger(__name__)

# PRODUCT Inv-17 (graceful-empty-with-recorded-reason, ratified S278).
# A structurally VALID workbook whose sheets match NEITHER the CSP nor the
# EFA scoring-matrix archetype yields zero fields. That is GRACEFUL (no raise
# — distinct from the strict-raise empty_xlsx / unreadable_xlsx paths) but it
# MUST NOT be silent: `extract` emits a structured log carrying this token and
# the downstream {52.12} form-write path threads it onto the `form_templates`
# row provenance so an `analysed`/0-field row always carries WHY. This is the
# single source of truth for that reason token (the test and any
# reason-threading consumer import it from here — no string duplication).
NO_ARCHETYPE_REASON = "no_archetype_match"


# ──────────────────────────────────────────────────────────────────────────
# Regex bank
# ──────────────────────────────────────────────────────────────────────────

# Inv-11 — word/page limit token. EFA's "Page Limit" column gives a
# bare integer; if non-numeric (e.g. ``Yes/No``), we leave word_limit
# unset. Inline ``[NNN] words`` is also recognised in question prose
# for parity with the PDF reader.
_WORD_LIMIT_BRACKETED = re.compile(r"\[(\d{1,5})\]\s*words?", re.IGNORECASE)
_WORD_LIMIT_PLAIN = re.compile(
    r"(?:max(?:imum)?\.?\s+|no more than\s+|up to\s+)?(\d{1,5})\s+words?",
    re.IGNORECASE,
)
_BARE_INT = re.compile(r"^\s*(\d{1,5})\s*$")

# EFA scoring-matrix Ref column: ``2.1``, ``3.1``, ``7.2`` etc.
_REF_PATTERN = re.compile(r"^\s*(\d+(?:\.\d+)*)\s*$")

# EFA scoring-matrix section-header row in the Criteria column —
# ``Part N - TITLE`` (case-insensitive on the dash, varies across rows).
_EFA_PART_HEADER = re.compile(
    r"^\s*(PART\s+\d+)\s*[\-–—:]\s*(.+?)\s*$",
    re.IGNORECASE,
)

# CSP letter key in column B — single uppercase A..Z OR ``PRINCIPLE  N``
# where the column carries an embedded newline (``"PRINCIPLE  \n1"``).
_CSP_LETTER_KEY = re.compile(r"^\s*([A-Z])\s*$")
_CSP_PRINCIPLE_KEY = re.compile(
    r"^\s*PRINCIPLE\s*\s*\n?\s*(\d+)\s*$",
    re.IGNORECASE | re.MULTILINE,
)

# Placeholder pattern — CSP's ``TYPE RESPONSE HERE>>>>`` shape plus the
# canonical PDF reader's placeholder vocabulary for parity.
_PLACEHOLDER_PATTERNS = [
    re.compile(r"^\s*TYPE\s+RESPONSE\s+HERE\s*>+\s*$", re.IGNORECASE),
    re.compile(r"^\[\s*insert\b.*?\]$", re.IGNORECASE),
    re.compile(r"^\[\s*enter\b.*?\]$", re.IGNORECASE),
    re.compile(r"^\[\s*type\b.*?\]$", re.IGNORECASE),
    re.compile(r"^\[\s*provide\b.*?\]$", re.IGNORECASE),
    re.compile(r"^\{\{[^}]+\}\}$"),
    re.compile(r"^<<[A-Z_ ]+>>$"),
    re.compile(r"^\{[A-Z_]+\}$"),
    re.compile(r"^n/?a$", re.IGNORECASE),
    re.compile(r"^-+$"),
    re.compile(r"^\.{3,}$"),
]

# URL pattern — fallback discovery for URLs in cell prose when openpyxl
# does not surface them as ``cell.hyperlink``.
_URL_PATTERN = re.compile(r"https?://[^\s\)\]\}]+", re.IGNORECASE)

# ──────────────────────────────────────────────────────────────────────────
# Generic (DR-058) detection — "any labelled cell with an adjacent empty
# answer cell -> a field" (TECH §3.2 / FORM-EXTRACTION-SPIKE §5), layered
# UNDER the CSP/EFA archetypes above as a fallback (only runs on a sheet
# neither archetype claimed). Header-span + minimum-repetition gated: a
# bare "any empty neighbour" scan over a whole free-form sheet produces
# real false positives — MEASURED on the EFA workbook's non-scoring-matrix
# sheets ("Title Sheet" metadata block, "Summary" sub-header rows) — so the
# fallback first locates a plausible header span (>=2 adjacent short,
# non-numeric cells) and only accepts the region below it once >= 3
# genuine label+empty-answer rows repeat (annex_3's rate card clears this
# easily at 10+; the EFA metadata sheets' 0-2 coincidental hits do not).
# ──────────────────────────────────────────────────────────────────────────

_GENERIC_HEADER_CELL_MAX_LEN = 60
_GENERIC_LABEL_MIN_LEN = 3
_GENERIC_MIN_REGION_MATCHES = 3
_GENERIC_MAX_BLANK_ROW_GAP = 3

_GENERIC_ANSWER_MARKER = re.compile(r"^[□☐○❑❒✓✗xX/-]{1,2}$")


def _looks_like_pure_number(text: str) -> bool:
    """True when text parses as a bare number (int/float, optional comma
    thousands-separators) — excluded from header-span candidacy (a numeric
    cell is data, never a column header)."""
    stripped = text.strip()
    if not stripped:
        return False
    try:
        float(stripped.replace(",", ""))
    except ValueError:
        return False
    return True


def _is_bare_zero(text: str) -> bool:
    """True when text is a bare numeric zero (``0``, ``0.0``, ``0.00``…) —
    the common "unfilled numeric input" default (measured on annex_3's
    rate-card Day Rate column). Distinct from ``_looks_like_pure_number``,
    which flags ANY number (used to keep a real number off the LABEL side
    of a generic pair, not to decide answer-shapedness)."""
    stripped = text.strip()
    if not stripped:
        return False
    try:
        return float(stripped) == 0.0
    except ValueError:
        return False


def _is_generic_answer_shaped(text: str) -> bool:
    """True when cell text reads as an unfilled answer slot for shape 3:
    empty, one of the archetype placeholder patterns, a short
    checkbox/tick-style marker glyph, or a bare numeric zero (an unfilled
    numeric-input default)."""
    if not text:
        return True
    if _is_placeholder(text):
        return True
    if _GENERIC_ANSWER_MARKER.fullmatch(text):
        return True
    return _is_bare_zero(text)


def _generic_header_spans(ws: Worksheet) -> list[tuple[int, int, int]]:
    """Find candidate generic header rows as ``(row, start_col, end_col)``
    for every maximal run of >= 2 adjacent short, non-numeric, non-empty
    cells. A sheet may carry several spans (e.g. two side-by-side
    sub-tables, as annex_3's "Rate Card" + "Resources" blocks)."""
    spans: list[tuple[int, int, int]] = []
    for row in range(1, ws.max_row + 1):
        run_start: int | None = None
        for col in range(1, ws.max_column + 2):  # +1 sentinel closes a trailing run
            text = (
                _clean(ws.cell(row=row, column=col).value)
                if col <= ws.max_column
                else ""
            )
            header_ish = (
                bool(text)
                and len(text) <= _GENERIC_HEADER_CELL_MAX_LEN
                and not _looks_like_pure_number(text)
            )
            if header_ish:
                if run_start is None:
                    run_start = col
            else:
                if run_start is not None and col - run_start >= 2:
                    spans.append((row, run_start, col - 1))
                run_start = None
    return spans


def _walk_generic_region(
    ws: Worksheet,
    *,
    header_row: int,
    start_col: int,
    end_col: int,
    max_row: int,
    table_index: int,
    sequence_start: int,
) -> tuple[list[ExtractedField], int]:
    """Walk rows below a generic header span, pairing adjacent
    labelled/empty-answer cells WITHIN the header's column window.

    Discards the whole region (returns no fields) when fewer than
    ``_GENERIC_MIN_REGION_MATCHES`` candidate rows are found — a lone
    coincidental match is far more likely a metadata-block artifact than
    a genuine repeating fillable table (measured false positives above).
    """
    candidates: list[ExtractedField] = []
    sequence = sequence_start
    blank_run = 0
    row = header_row + 1
    while row <= max_row:
        row_texts = {
            col: _clean(ws.cell(row=row, column=col).value)
            for col in range(start_col, end_col + 1)
        }
        if not any(row_texts.values()):
            blank_run += 1
            if blank_run >= _GENERIC_MAX_BLANK_ROW_GAP:
                break
            row += 1
            continue
        blank_run = 0
        consumed: set[int] = set()
        for col in range(start_col, end_col):
            if col in consumed or (col + 1) in consumed:
                continue
            left_text, right_text = row_texts[col], row_texts[col + 1]
            left_answer = _is_generic_answer_shaped(left_text)
            right_answer = _is_generic_answer_shaped(right_text)
            if left_answer == right_answer:
                continue
            if left_answer:
                label_col, label_text = col + 1, right_text
                answer_col, answer_text = col, left_text
            else:
                label_col, label_text = col, left_text
                answer_col, answer_text = col + 1, right_text
            if len(label_text) < _GENERIC_LABEL_MIN_LEN:
                continue
            if _looks_like_pure_number(label_text):
                # A bare number can never be a meaningful question label —
                # guards against a numeric data column (e.g. a "Weighted
                # Score" 0-default) masquerading as the label side once
                # bare-zero counts as answer-shaped (measured false
                # positive: EFA "Summary" sheet).
                continue
            is_ph = bool(answer_text) and _is_placeholder(answer_text)
            field_type = "placeholder" if is_ph else "empty_cell"
            candidates.append(
                ExtractedField(
                    question_text=label_text,
                    placeholder_text=answer_text if is_ph else None,
                    field_type=field_type,  # type: ignore[arg-type]
                    fill_status="pending",
                    row_index=row,
                    col_index=answer_col,
                    table_index=table_index,
                    section_name=None,
                    sequence=sequence,
                    word_limit=_extract_word_limit(label_text),
                    is_mandatory=None,
                    reference_urls=_collect_cell_hyperlinks(ws, row, answer_col),
                )
            )
            sequence += 1
            consumed.add(label_col)
            consumed.add(answer_col)
        row += 1

    if len(candidates) < _GENERIC_MIN_REGION_MATCHES:
        return [], sequence_start
    return candidates, sequence


def _emit_generic_sheet_fields(
    ws: Worksheet,
    *,
    table_index_offset: int,
    sequence_start: int,
) -> tuple[list[ExtractedField], int, int]:
    """Shape 3 — generic label -> empty-answer-cell fallback for a whole
    sheet: locate every candidate header span, walk the region below each
    (bounded by the next span or a blank-row gap), keep only regions that
    clear the repetition floor."""
    spans = _generic_header_spans(ws)
    if not spans:
        return [], sequence_start, 0

    all_fields: list[ExtractedField] = []
    sequence = sequence_start
    table_index = table_index_offset
    tables_consumed = 0
    for span_index, (header_row, start_col, end_col) in enumerate(spans):
        region_end = (
            spans[span_index + 1][0] - 1
            if span_index + 1 < len(spans)
            else ws.max_row
        )
        fields, sequence = _walk_generic_region(
            ws,
            header_row=header_row,
            start_col=start_col,
            end_col=end_col,
            max_row=region_end,
            table_index=table_index,
            sequence_start=sequence,
        )
        if fields:
            all_fields.extend(fields)
            table_index += 1
            tables_consumed += 1
    return all_fields, sequence, tables_consumed


# ──────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────


def _clean(value: Any) -> str:
    """Normalise a cell value to a single-line string.

    openpyxl returns ``None`` for blank cells; mixed types (int/float)
    for numeric cells; strings (possibly with embedded ``\\n``) for
    text. We coerce to ``str`` and collapse internal whitespace runs
    while preserving the trimmed value — substring matches survive
    line-wraps that way (e.g. ``"PRINCIPLE\\n1"`` → ``"PRINCIPLE 1"``).
    """
    if value is None:
        return ""
    if isinstance(value, str):
        flat = re.sub(r"\s+", " ", value).strip()
        return flat
    return str(value).strip()


def _raw_text(value: Any) -> str:
    """Return cell value as a string preserving original line breaks.

    Used where we want to detect line-anchored sub-keys (e.g. CSP's
    ``B1/`` prefix at the start of column C). Returns empty string for
    None / non-string values.
    """
    if value is None:
        return ""
    return str(value)


def _is_placeholder(text: str) -> bool:
    """Return True if the cell text is a placeholder scaffold."""
    stripped = text.strip()
    if not stripped:
        return False
    for pattern in _PLACEHOLDER_PATTERNS:
        if pattern.fullmatch(stripped):
            return True
    return False


def _extract_word_limit(text: str) -> int | None:
    """Inv-11 — extract a word/page limit from inline tokens.

    Used on the question prose (column C in CSP, column B in EFA
    scoring matrix). The EFA ``Page Limit`` column is handled
    separately via :func:`_parse_page_limit_cell`.
    """
    bracket = _WORD_LIMIT_BRACKETED.search(text)
    if bracket:
        return int(bracket.group(1))
    plain = _WORD_LIMIT_PLAIN.search(text)
    if plain:
        return int(plain.group(1))
    return None


def _parse_page_limit_cell(value: Any) -> int | None:
    """Parse the EFA ``Page Limit`` column's value into a word_limit int.

    The column is integer-typed when populated; ``Yes/No`` or other
    free-text values leave it None (Inv-11 — no inferred default).
    """
    if value is None:
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else None
    text = str(value).strip()
    if not text:
        return None
    bare = _BARE_INT.match(text)
    if bare:
        return int(bare.group(1))
    return None


def _collect_cell_hyperlinks(ws: Worksheet, row: int, col: int) -> list[str]:
    """Harvest URL targets from a cell's hyperlink + fallback regex on prose.

    openpyxl exposes ``cell.hyperlink.target`` for explicit links. We
    also re-scan the cell's text content for raw URLs (some XLSX files
    encode the URL as text rather than as a hyperlink object).
    """
    urls: list[str] = []
    cell = ws.cell(row=row, column=col)
    link = getattr(cell, "hyperlink", None)
    if link and getattr(link, "target", None):
        target = link.target
        if isinstance(target, str) and target.startswith(("http://", "https://")):
            urls.append(target)
    text = _raw_text(cell.value)
    for match in _URL_PATTERN.finditer(text):
        url = match.group(0).rstrip(".,;:")
        if url not in urls:
            urls.append(url)
    return urls


def _normalise_for_dedup(text: str | None) -> str:
    """Inv-13 dedup key — lowercase, collapse whitespace, strip terminal
    punctuation. Matches TECH §2.3's ``normalise`` definition."""
    if not text:
        return ""
    flat = re.sub(r"\s+", " ", text).strip().lower()
    flat = re.sub(r"[\.\,\;\:\!\?]+$", "", flat)
    return flat


# ──────────────────────────────────────────────────────────────────────────
# Per-sheet walkers
# ──────────────────────────────────────────────────────────────────────────


def _detect_efa_scoring_matrix_header(ws: Worksheet) -> int | None:
    """Return the row index of the EFA scoring-matrix header row, or None.

    The header row is the one carrying ``Ref`` in column 1 and
    ``Criteria`` in column 2. Used to drive the table walker — rows
    above the header are workbook metadata; rows below are
    section-header rows + question rows.
    """
    for row in range(1, ws.max_row + 1):
        ref = _clean(ws.cell(row=row, column=1).value)
        criteria = _clean(ws.cell(row=row, column=2).value)
        if ref.lower() == "ref" and criteria.lower() == "criteria":
            return row
    return None


def _walk_efa_scoring_matrix(
    ws: Worksheet,
    *,
    table_index: int,
    sequence_start: int,
) -> tuple[list[ExtractedField], int]:
    """Walk a sheet's EFA-style scoring matrix into ExtractedField rows.

    Returns the emitted field list + the next sequence counter.
    """
    header_row = _detect_efa_scoring_matrix_header(ws)
    if header_row is None:
        return [], sequence_start

    emitted: list[ExtractedField] = []
    sequence = sequence_start
    current_section: str | None = None

    # Find the Page Limit column (used by the compliance-check sub-table;
    # absent in the scoring matrix header). We re-scan the scoring matrix
    # header row to confirm — in this fixture it's absent, so this stays
    # None and we fall back to inline-text scanning for limits.
    page_limit_col: int | None = None
    for col in range(1, ws.max_column + 1):
        header_cell = _clean(ws.cell(row=header_row, column=col).value).lower()
        if "page limit" in header_cell:
            page_limit_col = col
            break

    for row in range(header_row + 1, ws.max_row + 1):
        ref = _clean(ws.cell(row=row, column=1).value)
        criteria_raw = ws.cell(row=row, column=2).value
        criteria = _clean(criteria_raw)
        if not ref and not criteria:
            continue

        # Section-header row: no Ref, Criteria carries ``Part N - TITLE``.
        if not ref and criteria:
            part_match = _EFA_PART_HEADER.match(criteria)
            if part_match:
                label = part_match.group(1).strip()
                descriptor = part_match.group(2).strip()
                current_section = f"{label} — {descriptor}"
                continue
            # Other section-style rows (``DESIGN TOTAL``, ``PRICING TOTAL``,
            # ``GRAND TOTAL``) — not questions; skip without emitting.
            continue

        # Question row: Ref is a numeric token + Criteria carries prose.
        if not _REF_PATTERN.match(ref):
            # E.g. score-meaning table at the bottom of the sheet
            # (column 2 carries a numeric score, column 3 a description).
            # These are reference rows, not questions — skip.
            continue
        if not criteria:
            # Numbered row with no question text — defensive skip.
            continue

        word_limit = None
        if page_limit_col is not None:
            word_limit = _parse_page_limit_cell(
                ws.cell(row=row, column=page_limit_col).value
            )
        if word_limit is None:
            word_limit = _extract_word_limit(criteria)

        emitted.append(
            ExtractedField(
                question_text=criteria,
                field_type="empty_cell",
                fill_status="pending",
                row_index=row,
                col_index=2,
                table_index=table_index,
                section_name=current_section,
                sequence=sequence,
                word_limit=word_limit,
                is_mandatory=None,
                reference_urls=_collect_cell_hyperlinks(ws, row, 2),
            )
        )
        sequence += 1

    return emitted, sequence


def _detect_csp_header(ws: Worksheet) -> int | None:
    """Return the row index of the CSP vendor-checklist header row.

    The header row has ``Principle`` in column 3 and ``Implementation
    Approach Used`` (or similar) in column 5.
    """
    for row in range(1, ws.max_row + 1):
        col3 = _clean(ws.cell(row=row, column=3).value).lower()
        col5 = _clean(ws.cell(row=row, column=5).value).lower()
        if col3.startswith("principle") and "implementation" in col5:
            return row
    return None


def _walk_csp_checklist(
    ws: Worksheet,
    *,
    table_index: int,
    sequence_start: int,
) -> tuple[list[ExtractedField], int]:
    """Walk a CSP vendor-checklist sheet into ExtractedField rows.

    The CSP structure:
      - rows 2-3: title-page metadata cells with ``TYPE RESPONSE HERE>>>>``
        placeholders. Emit as placeholder fields with the metadata
        label as ``question_text``.
      - row 5: header row.
      - rows 6-20: letter-keyed preamble (``A``, ``B``, ``C``, …).
        Column B has the letter key; column C has the section title +
        question prose (sometimes with ``B1/`` style sub-keys embedded).
        The response slot is column E.
      - rows 21-46: numbered principles (``PRINCIPLE N``). Column B has
        the key; column C the title + prose; column D the NCSC URL;
        column E the response slot. Sub-rows (``2.1``, ``9.1``, …) have
        no column B key and continue the prior principle.
    """
    header_row = _detect_csp_header(ws)
    if header_row is None:
        return [], sequence_start

    emitted: list[ExtractedField] = []
    sequence = sequence_start
    current_section: str | None = None

    # ── Title-page placeholders (rows above the header) ──
    for row in range(1, header_row):
        # Columns 3+5 carry label/response pairs in the fixture:
        # (label_col=3, response_col=4) and (label_col=5, response_col=6).
        for label_col, response_col in ((3, 4), (5, 6)):
            label = _clean(ws.cell(row=row, column=label_col).value)
            response = _clean(ws.cell(row=row, column=response_col).value)
            if not label or not response:
                continue
            if _is_placeholder(response):
                emitted.append(
                    ExtractedField(
                        question_text=label,
                        placeholder_text=response,
                        field_type="placeholder",
                        fill_status="pending",
                        row_index=row,
                        col_index=response_col,
                        table_index=table_index,
                        section_name="Document Header",
                        sequence=sequence,
                        word_limit=None,
                        is_mandatory=None,
                        reference_urls=[],
                    )
                )
                sequence += 1

    # ── Question rows (header_row + 1 .. max_row) ──
    for row in range(header_row + 1, ws.max_row + 1):
        key_raw = ws.cell(row=row, column=2).value
        key_clean = _clean(key_raw)
        prose_raw = ws.cell(row=row, column=3).value
        prose_clean = _clean(prose_raw)

        # Detect section markers in column B.
        if key_clean:
            principle_match = _CSP_PRINCIPLE_KEY.match(_raw_text(key_raw).strip())
            if principle_match:
                # ``PRINCIPLE N`` row — start a new section.
                principle_num = principle_match.group(1)
                # Section name: principle key plus the title from col C
                # (split on newline / colon to get the title).
                title = ""
                if prose_clean:
                    # Title is the first line of col C, stripped of leading
                    # numbers (``1 Data in transit protection`` →
                    # ``Data in transit protection``).
                    first_line = _raw_text(prose_raw).split("\n", 1)[0].strip()
                    title = re.sub(
                        r"^\s*(?:Principle\s+)?\d+\s*[:\.\-–—]?\s*",
                        "",
                        first_line,
                    ).strip()
                current_section = (
                    f"PRINCIPLE {principle_num} — {title}"
                    if title
                    else f"PRINCIPLE {principle_num}"
                )
                # Emit the principle row itself as a question row when its
                # response slot is a placeholder (Inv-9). The first line is
                # the title, the remainder is the prose.
                if _emit_csp_question(
                    ws=ws,
                    row=row,
                    table_index=table_index,
                    sequence=sequence,
                    section=current_section,
                    prose_raw=prose_raw,
                    emitted=emitted,
                ):
                    sequence += 1
                continue

            letter_match = _CSP_LETTER_KEY.match(key_clean)
            if letter_match:
                letter = letter_match.group(1)
                # Section name: letter key + title from col C first line.
                title = ""
                if prose_clean:
                    first_line = _raw_text(prose_raw).split("\n", 1)[0].strip()
                    title = first_line
                current_section = (
                    f"Section {letter} — {title}" if title else f"Section {letter}"
                )
                if _emit_csp_question(
                    ws=ws,
                    row=row,
                    table_index=table_index,
                    sequence=sequence,
                    section=current_section,
                    prose_raw=prose_raw,
                    emitted=emitted,
                ):
                    sequence += 1
                continue

        # Sub-row: no key in col B; inherits ``current_section``.
        if prose_clean:
            if _emit_csp_question(
                ws=ws,
                row=row,
                table_index=table_index,
                sequence=sequence,
                section=current_section,
                prose_raw=prose_raw,
                emitted=emitted,
            ):
                sequence += 1

    return emitted, sequence


def _emit_csp_question(
    *,
    ws: Worksheet,
    row: int,
    table_index: int,
    sequence: int,
    section: str | None,
    prose_raw: Any,
    emitted: list[ExtractedField],
) -> bool:
    """Emit one CSP question row, classifying placeholder vs authored.

    Inv-9: the question_text is the prose; the response slot (column 5)
    determines field_type. NCSC URLs (column 4 hyperlink) → reference_urls.

    Returns True when a field was appended (the caller advances the
    sequence counter); False when the row carried no prose.
    """
    prose = _clean(prose_raw)
    if not prose:
        return False

    response_value = ws.cell(row=row, column=5).value
    response_text = _clean(response_value)

    if _is_placeholder(response_text):
        field_type = "placeholder"
        placeholder_text = response_text
    else:
        field_type = "empty_cell"
        placeholder_text = None

    # NCSC URL is in column 4 of the principle row (and absent on sub-rows).
    # Also fall back to scanning the prose for raw URLs.
    reference_urls = _collect_cell_hyperlinks(ws, row, 4)
    # Add any URLs in the prose itself.
    for match in _URL_PATTERN.finditer(_raw_text(prose_raw)):
        url = match.group(0).rstrip(".,;:")
        if url not in reference_urls:
            reference_urls.append(url)

    emitted.append(
        ExtractedField(
            question_text=prose,
            placeholder_text=placeholder_text,
            field_type=field_type,
            fill_status="pending",
            row_index=row,
            col_index=3,
            table_index=table_index,
            section_name=section,
            sequence=sequence,
            word_limit=_extract_word_limit(prose),
            is_mandatory=None,
            reference_urls=reference_urls,
        )
    )
    return True


# ──────────────────────────────────────────────────────────────────────────
# Sheet dispatcher
# ──────────────────────────────────────────────────────────────────────────


def _walk_sheet(
    ws: Worksheet,
    *,
    table_index_offset: int,
    sequence_start: int,
) -> tuple[list[ExtractedField], int, int]:
    """Dispatch a sheet to the appropriate walker.

    Returns ``(fields, next_sequence, tables_consumed)``. The dispatcher
    tries each archetype in turn; the first that yields a header match
    is the one applied. A sheet matching neither archetype falls to the
    generic shape-3 fallback (DR-058 — "drop the EFA/CSP-only gate").
    """
    # Try CSP archetype first — it has a more specific header signature
    # (``Principle`` in col 3 + ``Implementation`` in col 5).
    if _detect_csp_header(ws) is not None:
        fields, next_seq = _walk_csp_checklist(
            ws,
            table_index=table_index_offset,
            sequence_start=sequence_start,
        )
        return fields, next_seq, 1

    # Try EFA scoring matrix.
    if _detect_efa_scoring_matrix_header(ws) is not None:
        fields, next_seq = _walk_efa_scoring_matrix(
            ws,
            table_index=table_index_offset,
            sequence_start=sequence_start,
        )
        return fields, next_seq, 1

    # Shape 3 (DR-058 generalisation): generic label -> empty-answer-cell
    # fallback, layered UNDER the two archetypes above.
    return _emit_generic_sheet_fields(
        ws,
        table_index_offset=table_index_offset,
        sequence_start=sequence_start,
    )


# ──────────────────────────────────────────────────────────────────────────
# Dedup (Inv-13 / TECH §2.3)
# ──────────────────────────────────────────────────────────────────────────


def _dedup_per_form(fields: list[ExtractedField]) -> list[ExtractedField]:
    """Inv-13 / TECH §2.3 — per-form dedup keyed on
    ``(section_name, normalise(question_text))``.

    Keeps the FIRST occurrence in reading order; discards subsequent
    copies. Preserves the original coordinates from the first
    occurrence. After dedup, ``sequence`` is recomputed as the new
    reading-order index so the contract that ``sequence`` is strictly
    increasing across the surviving rows is preserved.
    """
    seen: set[tuple[str | None, str]] = set()
    survivors: list[ExtractedField] = []
    for field in fields:
        if field.question_text is None:
            # Pure-placeholder rows (no authored question) are not
            # deduplicated against authored questions — they key on
            # placeholder_text + section.
            key = (field.section_name, _normalise_for_dedup(field.placeholder_text))
        else:
            key = (field.section_name, _normalise_for_dedup(field.question_text))
        if key in seen:
            continue
        seen.add(key)
        survivors.append(field)

    # Recompute sequence as the post-dedup reading-order index — drop the
    # gaps left by removed duplicates.
    return [
        field.model_copy(update={"sequence": i})
        for i, field in enumerate(survivors)
    ]


# ──────────────────────────────────────────────────────────────────────────
# Public entry point
# ──────────────────────────────────────────────────────────────────────────


async def extract(raw_bytes: bytes, filename: str) -> ExtractedForm:
    """Extract structured fields from a blank XLSX form.

    Args:
        raw_bytes: The XLSX file's bytes (read by the caller from the
            ingest source).
        filename: The file's name/path relative to the ingest source.
            Passed through into ``FormExtractionError.rel_path`` and
            used as a ``FormMetadata.form_title`` fallback when no
            sheet-level title is detectable.

    Returns:
        An ``ExtractedForm`` carrying the form-level metadata + per-field
        rows in reading order, after per-form dedup (Inv-13).

    Raises:
        FormExtractionError: When the XLSX cannot be opened or contains
            zero worksheets of readable content (Inv-17 — never silently
            returns an empty ``ExtractedForm`` when extraction itself
            failed).
    """
    if not raw_bytes:
        raise FormExtractionError(
            reason="empty_xlsx",
            rel_path=filename,
            details="received zero bytes",
        )

    try:
        wb = openpyxl.load_workbook(
            io.BytesIO(raw_bytes),
            data_only=True,
            read_only=False,
        )
    except Exception as exc:  # noqa: BLE001 — openpyxl surfaces a wide mix of
        # types on corrupt input (InvalidFileException, KeyError, OSError,
        # ValueError, …); every load failure maps to the same typed error.
        raise FormExtractionError(
            reason="unreadable_xlsx",
            rel_path=filename,
            details=str(exc),
        ) from exc

    if not wb.sheetnames:
        raise FormExtractionError(
            reason="empty_xlsx",
            rel_path=filename,
            details="workbook has zero sheets",
        )

    all_fields: list[ExtractedField] = []
    sequence = 0
    table_index = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        fields, sequence, consumed = _walk_sheet(
            ws,
            table_index_offset=table_index,
            sequence_start=sequence,
        )
        all_fields.extend(fields)
        table_index += consumed

    # Inv-13 — per-form dedup.
    deduped = _dedup_per_form(all_fields)

    # PRODUCT Inv-17 (graceful-empty-with-recorded-reason, ratified S278). A
    # structurally VALID workbook whose sheets matched no archetype yields zero
    # fields. This is GRACEFUL (we return an empty ``ExtractedForm`` rather than
    # raising — distinct from the strict-raise empty/unreadable/zero-sheet
    # paths above), but it MUST NOT be SILENT: surface a structured,
    # machine-readable reason so the {52.12} form-write path can RECORD it on
    # the ``form_templates`` row provenance. Without this, an ``analysed``/
    # 0-field row would carry no reason — the exact shape Inv-17 forbids.
    if not deduped:
        _logger.info(
            json.dumps(
                {
                    "event": "form_extractor.zero_archetype",
                    "reason": NO_ARCHETYPE_REASON,
                    "rel_path": filename,
                    "form_format": "xlsx",
                    "sheet_count": len(wb.sheetnames),
                }
            )
        )

    metadata = FormMetadata(
        form_type="questionnaire",
        form_format="xlsx",
        form_title=filename,
    )
    return ExtractedForm(form_metadata=metadata, fields=deduped)
