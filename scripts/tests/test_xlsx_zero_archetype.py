"""Inv-17 graceful-empty-with-recorded-reason — ratified semantic (S278).

Background
----------

The XLSX extractor (``scripts/cocoindex_pipeline/form_extractors/xlsx.py``)
``extract()`` RAISES ``FormExtractionError`` on the genuinely-broken inputs:

  - empty bytes (``reason='empty_xlsx'``),
  - unreadable / non-XLSX bytes (``reason='unreadable_xlsx'``),
  - a workbook with zero sheets (``reason='empty_xlsx'``).

Those strict paths are already covered by
``TestXlsxFailureSurfacing`` in ``test_form_extractors.py`` and STAY
strict per the ratified PRODUCT Inv-17 amendment.

There is a fourth case that is NOT a raise: a structurally VALID
workbook whose sheets match NEITHER the CSP archetype NOR the EFA
scoring-matrix archetype. ``_walk_sheet`` returns ``[]`` for such a
sheet (see the ``return [], sequence_start, 0`` fall-through and its
docstring line "Sheets matching no archetype yield no fields"), and
``extract()`` then returns ``ExtractedForm(fields=[])``.

Disposition (ratified S278)
---------------------------

Liam ratified the GRACEFUL disposition for the zero-archetype-but-valid
workbook: ``extract`` STAYS graceful (returns an empty ``ExtractedForm``,
NOT a strict raise) — DISTINCT from the strict-raise paths above. But the
graceful path MUST NOT be SILENT: ``extract`` emits a structured,
surfaced log carrying a machine-readable no-archetype reason, so the
downstream {52.12} form-write path can thread that reason onto the
``form_templates`` row provenance. The shape PRODUCT Inv-17 forbids
("an instance record but silently zero fields with no recorded reason")
is closed: the reason is now both LOGGED (surfaced) and RECORDED (row
provenance — asserted in ``test_cocoindex_flow_write_path.py``).

This test therefore asserts the GRACEFUL-WITH-RECORDED-REASON semantic:
empty form returned (no raise) AND a structured ``form_extractor``
zero-archetype log emitted carrying the reason. If a future change
either (a) silences the log or (b) flips to a strict raise, this test
fails loudly and the author is routed back to the ratified Inv-17
amendment.

References:
  - docs/specs/id-52-form-extraction/PRODUCT.md Inv-17 (graceful-empty
    -with-recorded-reason admitted as a valid shape, S278).
  - scripts/cocoindex_pipeline/form_extractors/xlsx.py
    (``NO_ARCHETYPE_REASON``; ``extract`` emits the structured log and
     returns ``ExtractedForm(fields=[])``).
  - docs/reference/testing/test-philosophy.md (real-behaviour: builds a real
    workbook with openpyxl, runs the real extractor — no mocks).
"""

from __future__ import annotations

import asyncio
import io
import json
import logging

import openpyxl

from scripts.cocoindex_pipeline.form_extractors.shared import ExtractedForm
from scripts.cocoindex_pipeline.form_extractors.xlsx import (
    NO_ARCHETYPE_REASON,
    extract as xlsx_extract,
)


def _build_zero_archetype_workbook() -> bytes:
    """Construct a structurally VALID XLSX whose single sheet matches
    NEITHER the CSP archetype (``Principle`` in col 3 + ``Implementation``
    in col 5) NOR the EFA scoring-matrix archetype (``Ref`` in col 1 +
    ``Criteria`` in col 2).

    The content is plain freeform prose, so ``_detect_csp_header`` and
    ``_detect_efa_scoring_matrix_header`` both return ``None`` and the
    sheet yields zero fields — but the workbook itself is well-formed
    (openpyxl opens it cleanly, ``sheetnames`` is non-empty), so the
    strict-raise guards in ``extract`` are NOT triggered.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Project Notes"
    # Deliberately avoid both archetype header signatures.
    ws["A1"] = "Internal Project Notes"
    ws["A2"] = "This worksheet is freeform content, not a structured form."
    ws["B2"] = "No 'Ref'/'Criteria' header; no 'Principle'/'Implementation' header."
    ws["A3"] = "Owner"
    ws["B3"] = "Procurement Team"
    ws["A4"] = "Status"
    ws["B4"] = "Draft"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestZeroArchetypeXlsxGracefulSemantic:
    """Inv-17 graceful-empty-with-recorded-reason — ratified semantic (S278)."""

    def test_zero_archetype_xlsx_returns_empty_form_not_error(self) -> None:
        """Ratified GRACEFUL disposition (S278): a structurally VALID XLSX whose
        sheets match no archetype returns an ``ExtractedForm`` with
        ``fields == []`` and DOES NOT raise ``FormExtractionError``.

        This is DISTINCT from the strict-raise paths (empty bytes / unreadable
        bytes / zero-sheet workbook), which stay strict.
        """
        raw = _build_zero_archetype_workbook()

        form = asyncio.run(xlsx_extract(raw, "zero-archetype.xlsx"))

        assert isinstance(form, ExtractedForm)
        assert form.fields == []
        assert len(form.fields) == 0
        # The form-level metadata still constructs (filename fallback title).
        assert form.form_metadata.form_title == "zero-archetype.xlsx"

    def test_zero_archetype_xlsx_emits_surfaced_reason_not_silent(
        self,
        caplog,
    ) -> None:
        """Inv-17 graceful-empty MUST NOT be silent: ``extract`` emits a
        structured ``form_extractor`` log carrying the machine-readable
        no-archetype reason, so the downstream form-write path can thread a
        RECORDED reason onto the ``form_templates`` row.

        This is the forcing function that closes the "silently zero fields
        with no recorded reason" shape PRODUCT Inv-17 forbids: the reason is
        SURFACED here, and RECORDED on the row (see
        ``test_cocoindex_flow_write_path.py``)."""
        raw = _build_zero_archetype_workbook()

        with caplog.at_level(logging.INFO):
            asyncio.run(xlsx_extract(raw, "zero-archetype.xlsx"))

        # Locate the structured zero-archetype log line emitted by `extract`.
        structured_records: list[dict[str, object]] = []
        for record in caplog.records:
            try:
                payload = json.loads(record.getMessage())
            except (ValueError, TypeError):
                continue
            if isinstance(payload, dict) and payload.get("reason") == NO_ARCHETYPE_REASON:
                structured_records.append(payload)

        assert structured_records, (
            "expected a structured form_extractor log carrying "
            f"reason={NO_ARCHETYPE_REASON!r}; got "
            f"{[r.getMessage() for r in caplog.records]!r}"
        )
        payload = structured_records[0]
        # The surfaced reason must identify the file and the event so the
        # operator (and the downstream write path) can correlate it.
        assert payload["reason"] == NO_ARCHETYPE_REASON
        assert payload["rel_path"] == "zero-archetype.xlsx"
        assert "event" in payload

    def test_no_archetype_reason_is_a_stable_machine_readable_token(self) -> None:
        """``NO_ARCHETYPE_REASON`` is the single source of truth for the
        graceful-empty reason token — a stable, lowercase, machine-readable
        string both the extractor log and the row-provenance thread share."""
        assert isinstance(NO_ARCHETYPE_REASON, str)
        assert NO_ARCHETYPE_REASON
        assert NO_ARCHETYPE_REASON == NO_ARCHETYPE_REASON.strip()
        assert " " not in NO_ARCHETYPE_REASON

    def test_workbook_is_genuinely_valid_not_the_strict_raise_path(self) -> None:
        """Guard: prove the fixture is a VALID workbook (openpyxl opens it,
        has >= 1 sheet) so the zero-field outcome is the archetype-miss path,
        NOT one of the already-covered strict-raise paths (empty bytes /
        unreadable bytes / zero-sheet workbook)."""
        raw = _build_zero_archetype_workbook()
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=False)
        assert wb.sheetnames  # non-empty → not the zero-sheet raise path
        assert raw  # non-empty bytes → not the empty_xlsx raise path
